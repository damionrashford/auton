"""Multi-format document parser.

Extracts text content from various file types using lightweight,
targeted libraries:

  - PDF: pymupdf (fitz) — ultra-fast C library
  - Word (.docx): python-docx
  - Excel (.xlsx): openpyxl
  - CSV: stdlib csv module
  - Jupyter (.ipynb): nbformat
  - Images: pytesseract + Pillow (OCR)
  - Markdown/Text/Code: built-in

Each parser returns a list of ``PageContent`` objects containing the
extracted text and metadata (page number, sheet name, etc.).
"""

from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


@dataclass
class PageContent:
    """A single page/section of extracted text from a document."""

    text: str
    page_number: int = 0
    section: str = ""
    metadata: dict[str, Any] = field(default_factory=dict)


# File extensions mapped to parser functions.
_PARSER_MAP: dict[str, str] = {
    # PDF
    ".pdf": "pdf",
    # Word
    ".docx": "docx",
    ".doc": "docx",
    # Excel
    ".xlsx": "xlsx",
    ".xls": "xlsx",
    # CSV
    ".csv": "csv",
    ".tsv": "csv",
    # Jupyter
    ".ipynb": "notebook",
    # Images (OCR)
    ".png": "image",
    ".jpg": "image",
    ".jpeg": "image",
    ".gif": "image",
    ".bmp": "image",
    ".tiff": "image",
    ".webp": "image",
    # Text-based (direct read)
    ".txt": "text",
    ".md": "text",
    ".markdown": "text",
    ".rst": "text",
    ".json": "text",
    ".yaml": "text",
    ".yml": "text",
    ".toml": "text",
    ".xml": "text",
    ".html": "text",
    ".htm": "text",
    # Code files
    ".py": "text",
    ".js": "text",
    ".ts": "text",
    ".jsx": "text",
    ".tsx": "text",
    ".java": "text",
    ".go": "text",
    ".rs": "text",
    ".rb": "text",
    ".php": "text",
    ".c": "text",
    ".cpp": "text",
    ".h": "text",
    ".cs": "text",
    ".swift": "text",
    ".kt": "text",
    ".sql": "text",
    ".sh": "text",
    ".bash": "text",
    ".zsh": "text",
    ".r": "text",
    ".scala": "text",
    ".lua": "text",
    ".css": "text",
    ".scss": "text",
    ".less": "text",
    ".graphql": "text",
    ".proto": "text",
    ".tf": "text",
    ".dockerfile": "text",
}


def parse_document(file_path: str, content: bytes | None = None) -> list[PageContent]:
    """Parse a document file into a list of page contents.

    Args:
        file_path: Path or filename (used for extension detection).
        content: Raw file bytes. If None, reads from file_path.

    Returns:
        List of PageContent objects with extracted text.
    """
    ext = Path(file_path).suffix.lower()
    parser_type = _PARSER_MAP.get(ext, "text")

    if content is None:
        content = Path(file_path).read_bytes()

    try:
        if parser_type == "pdf":
            return _parse_pdf(content, file_path)
        if parser_type == "docx":
            return _parse_docx(content, file_path)
        if parser_type == "xlsx":
            return _parse_xlsx(content, file_path)
        if parser_type == "csv":
            return _parse_csv(content, file_path)
        if parser_type == "notebook":
            return _parse_notebook(content, file_path)
        if parser_type == "image":
            return _parse_image(content, file_path)
        # Default: plain text
        return _parse_text(content, file_path)
    except Exception:
        logger.warning("Failed to parse %s as %s", file_path, parser_type, exc_info=True)
        # Fallback: try as plain text
        try:
            return _parse_text(content, file_path)
        except Exception:
            return [PageContent(text=f"[error] Could not parse file: {file_path}")]


def _parse_pdf(content: bytes, file_path: str) -> list[PageContent]:
    """Extract text from PDF using pymupdf."""
    import fitz  # type: ignore[import-untyped]  # pymupdf

    pages: list[PageContent] = []
    with fitz.open(stream=content, filetype="pdf") as doc:
        for i, page in enumerate(doc):
            text: str = page.get_text("text")
            if text.strip():
                pages.append(
                    PageContent(
                        text=text.strip(),
                        page_number=i + 1,
                        metadata={"source": file_path},
                    )
                )
    return pages


def _parse_docx(content: bytes, file_path: str) -> list[PageContent]:
    """Extract text from Word documents."""
    from docx import Document

    doc = Document(io.BytesIO(content))
    paragraphs: list[str] = []
    for para in doc.paragraphs:
        para_text: str = para.text.strip()
        if para_text:
            paragraphs.append(para_text)

    if not paragraphs:
        return []

    # Group into logical pages (~3000 chars each)
    pages: list[PageContent] = []
    current: list[str] = []
    current_len = 0
    page_num = 1

    for p_text in paragraphs:
        current.append(p_text)
        current_len += len(p_text)
        if current_len >= 3000:
            pages.append(
                PageContent(
                    text="\n\n".join(current),
                    page_number=page_num,
                    metadata={"source": file_path},
                )
            )
            current = []
            current_len = 0
            page_num += 1

    if current:
        pages.append(
            PageContent(
                text="\n\n".join(current),
                page_number=page_num,
                metadata={"source": file_path},
            )
        )
    return pages


def _parse_xlsx(content: bytes, file_path: str) -> list[PageContent]:
    """Extract text from Excel spreadsheets."""
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    pages: list[PageContent] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append(" | ".join(cells))

        if rows:
            # Include header as context
            text = f"Sheet: {sheet_name}\n" + "\n".join(rows)
            pages.append(
                PageContent(
                    text=text,
                    section=sheet_name,
                    metadata={"source": file_path, "sheet": sheet_name},
                )
            )

    wb.close()
    return pages


def _parse_csv(content: bytes, file_path: str) -> list[PageContent]:
    """Parse CSV/TSV files into structured text."""
    text = content.decode("utf-8", errors="replace")
    delimiter = "\t" if file_path.endswith(".tsv") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [" | ".join(row) for row in reader if any(row)]
    if not rows:
        return []
    return [
        PageContent(
            text="\n".join(rows),
            metadata={"source": file_path},
        )
    ]


def _parse_notebook(content: bytes, file_path: str) -> list[PageContent]:
    """Extract code and markdown cells from Jupyter notebooks."""
    import nbformat

    nb = nbformat.reads(content.decode("utf-8"), as_version=4)  # type: ignore[no-untyped-call]
    pages: list[PageContent] = []

    for i, cell in enumerate(nb.cells):
        cell_type = cell.cell_type
        source = cell.source.strip()
        if not source:
            continue

        prefix = f"[{cell_type} cell {i + 1}]"
        pages.append(
            PageContent(
                text=f"{prefix}\n{source}",
                page_number=i + 1,
                section=cell_type,
                metadata={"source": file_path, "cell_type": cell_type},
            )
        )

    return pages


def _parse_image(content: bytes, file_path: str) -> list[PageContent]:
    """Extract text from images using OCR (pytesseract)."""
    try:
        import pytesseract  # type: ignore[import-not-found]
        from PIL import Image  # type: ignore[import-not-found]

        img = Image.open(io.BytesIO(content))
        text = pytesseract.image_to_string(img)
        if text.strip():
            return [
                PageContent(
                    text=text.strip(),
                    metadata={"source": file_path, "ocr": True},
                )
            ]
        return [PageContent(text="[no text detected in image]")]
    except ImportError:
        logger.warning(
            "pytesseract or Pillow not installed. "
            "Image OCR not available."
        )
        return [PageContent(text="[image OCR not available — install pytesseract]")]


def _parse_text(content: bytes, file_path: str) -> list[PageContent]:
    """Parse plain text, markdown, code, and other text files."""
    text = content.decode("utf-8", errors="replace").strip()
    if not text:
        return []
    return [
        PageContent(
            text=text,
            metadata={"source": file_path},
        )
    ]


def supported_extensions() -> list[str]:
    """Return all supported file extensions."""
    return sorted(_PARSER_MAP.keys())
